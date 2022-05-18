import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl import *

print("Starting cleanup of Kibana reports ...")

report_one = './.data/CAMPD - Prod APP ERROR Checks.csv'

report_two = './.data/CAMPD - Prod CELL DESTROYED Audit.csv'

report_three = './.data/CAMPD - Prod STG Checks.csv'

report_four = './.data/CAMPD - Prod RTR Checks.csv'

# Cleanup of CAMPD - Prod APP ERROR Checks
print("Starting Cleanup of CAMPD - Prod APP ERROR Checks")

# Covert csv to xlsx
read_one = pd.read_csv (report_one)
new_report_one = pd.ExcelWriter('./.data/CAMPD-Prod-APP-ERROR-Checks.xlsx')
read_one.to_excel (new_report_one, index = None, header=True)

new_report_one.save()

# Add filter to columns 
wb_one = load_workbook(filename=new_report_one)
ws_one = wb_one.active

ws_one.auto_filter.ref  = ws_one.dimensions
wb_one.save('./.data/final-CAMPD-Prod-APP-ERROR-Checks.xlsx')

print("Completed Cleanup of CAMPD - Prod APP ERROR Checks")

# Cleanup of CAMPD - Prod CELL DESTROYED Audit
print("Starting Cleanup of CAMPD - Prod CELL DESTROYED Audit")

# Covert csv to xlsx
read_two = pd.read_csv (report_two)
new_report_two = pd.ExcelWriter('./.data/CAMPD-Prod-CELL-DESTROYED-Audit.xlsx')
read_two.to_excel (new_report_two, index = None, header=True)

new_report_two.save()

# Add filter to columns 
wb_two = load_workbook(filename=new_report_two)
ws_two = wb_two.active

ws_two.auto_filter.ref  = ws_two.dimensions
wb_two.save('./.data/final-CAMPD-Prod-CELL-DESTROYED-Audit.xlsx')

print("Completed Cleanup of CAMPD - Prod CELL DESTROYED Audit")



# Cleanup of CAMPD - Prod STG Checks
print("Starting Cleanup of CAMPD - Prod STG Checks")

# Covert csv to xlsx
read_three = pd.read_csv (report_three)
new_report_three = pd.ExcelWriter('./.data/CAMPD-Prod-STG-Checks.xlsx')
read_three.to_excel (new_report_three, index = None, header=True)

new_report_three.save()

# Add filter to columns 
wb_three = load_workbook(filename=new_report_three)
ws_three = wb_three.active

ws_three.auto_filter.ref  = ws_three.dimensions
wb_three.save('./.data/final-CAMPD-Prod-STG-Checks.xlsx')

print("Completed Cleanup of CAMPD - Prod STG Checks")


# Cleanup of CAMPD - Prod RTR Checks
print("Starting Cleanup of CAMPD - Prod RTR Checks")

# Covert csv to xlsx
read_four = pd.read_csv (report_four)
new_report_four = pd.ExcelWriter('./.data/CAMPD-Prod-RTR-Checks.xlsx')
read_four.to_excel (new_report_four, index = None, header=True)

new_report_four.save()

# Add filter to columns 
wb_four = load_workbook(filename=new_report_four)
ws_four = wb_four.active

ws_four.auto_filter.ref  = ws_four.dimensions
wb_four.save('./.data/final-CAMPD-Prod-RTR-Checks.xlsx')

print("Completed Cleanup of CAMPD - Prod RTR Checks")